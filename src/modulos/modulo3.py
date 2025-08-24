import sqlite3
import customtkinter
from tkinter import ttk, messagebox, filedialog
from util.db import *
import openpyxl
from openpyxl.styles import Font, Alignment
from colores import *


class Modulo3(customtkinter.CTkToplevel):
    def __init__(self, master=None):
        super().__init__(master)
        self.title("Perfil Topográfico")
        self.geometry("300x200")
        label = customtkinter.CTkLabel(self, text='Gestión de Estaciones', width=40, height=28, fg_color='transparent', font=("Arial", 24, "bold"))
        label.pack(pady=10)

        # SECCIONES
        self.frame_izquierdo = customtkinter.CTkFrame(self, fg_color=color_fondo2, border_width=1, border_color="#221E1E", width=200)
        self.frame_izquierdo.pack(side="left", fill="y", padx=10, pady=5)
        self.frame_izquierdo.pack_propagate(False)

        self.frame_derecho = customtkinter.CTkFrame(self, fg_color=color_fondo2, border_width=1, border_color="#221E1E")
        self.frame_derecho.pack(side="left", fill="both", expand=True, pady=5, padx=10)

        # ETIQUETAS E INPUT LADO IZQUIERDO        
        frame_botones = customtkinter.CTkFrame(self.frame_izquierdo, fg_color="transparent")
        frame_botones.pack(pady=20, padx=20, fill="both", expand=True)
        boton_crud_1 = customtkinter.CTkButton(frame_botones, text="+ Agregar Estación", command=self.agregar, fg_color="#4c607c")
        boton_crud_1.pack(side="top", pady=(65, 15), fill="x")
        boton_crud_2 = customtkinter.CTkButton(frame_botones, text="Actualizar Estación", command=self.actualizar, fg_color="#4c607c")
        boton_crud_2.pack(side="top", pady=(15, 15), fill="x")
        boton_crud_3 = customtkinter.CTkButton(frame_botones, text="Eliminar Estación", command=self.eliminar, fg_color="#4c607c")
        boton_crud_3.pack(side="top", pady=(15, 15), fill="x")
        boton_crud_4 = customtkinter.CTkButton(frame_botones, text="Limpiar Datos", command=self.limpiar_campos, fg_color="#4c607c")
        boton_crud_4.pack(side="top", pady=10, fill="x")

        # LINEA SEPARADORA
        separator = customtkinter.CTkFrame(frame_botones, height=2, fg_color="#5E5757")
        separator.pack(fill="x", pady=10)

        boton1 = customtkinter.CTkButton(frame_botones, text="Importar Datos", command=self.importar_desde_excel, fg_color="#4c607c")
        boton1.pack(side="top", pady=15, fill="x")
        boton2 = customtkinter.CTkButton(frame_botones, text="Exportar Datos", command=self.exportar_a_excel, fg_color="#4c607c")
        boton2.pack(side="top", pady=15, fill="x")
        boton3 = customtkinter.CTkButton(frame_botones, text="Vaciar Base Datos", command=self.vaciar_base_datos, fg_color="#4c607c")
        boton3.pack(side="top", pady=15, fill="x")
        boton4 = customtkinter.CTkButton(frame_botones, text="Cerrar Módulo", command=self.cerrar_modulo, fg_color="#4c607c")
        boton4.pack(side="top", pady=15, fill="x")

        # DESARROLLO DEL CRUD

        # CAMPOS FORMULARIO
        self.campos = {}
        labels = [
            "Nombre estación", "Latitud", "Longitud", "Altura torre",
            "Estado", "Municipio", "Parroquia", "Tipo estación", "Observaciones"
        ]
        for i, texto in enumerate(labels):
            lbl = customtkinter.CTkLabel(self, text=texto+":", fg_color='transparent', font=("Arial", 14, "bold"))
            lbl.place(x=260, y=100 + i*40)
            entry = customtkinter.CTkEntry(self, width=200)
            entry.place(x=400, y=100 + i*40)
            self.campos[texto] = entry

        # TABLA
        style = ttk.Style(self)
        style.theme_use("clam")
        bg_color = color_fondo1        # fondo filas
        fg_color = "white"          # texto normal
        selected_bg = "#347083"     # fondo fila seleccionada
        selected_fg = "white"       # texto fila seleccionada
        heading_bg = color_fondo3      # fondo encabezados
        heading_fg = "white"        # texto encabezados
        field_bg = color_fondo1        # fondo área datos
        style.configure("Custom.Treeview",
                        background=bg_color,
                        foreground=fg_color,
                        fieldbackground=field_bg,
                        bordercolor=bg_color,
                        borderwidth=0,
                        font=("Arial", 10))
        style.configure("Custom.Treeview.Heading",
                        background=heading_bg,
                        foreground=heading_fg,
                        font=("Arial", 10))

        style.map("Custom.Treeview",
                  background=[('selected', selected_bg)],
                  foreground=[('selected', selected_fg)])
        

        columnas = ("id", "nombre_estacion", "latitud", "longitud", "altura_torre", "estado")
        self.tree = ttk.Treeview(self.frame_derecho, columns=columnas, show="headings", style="Custom.Treeview")     

        # CAMPO PARA HACER FILTRO TEMPORAL
        self.label_buscartexto1 = customtkinter.CTkLabel(self.frame_derecho, text="Filtrar:", font=("Arial", 12, "bold"))
        self.label_buscartexto1.place(x=400, y=640)
        self.punto_buscarpropuesto = customtkinter.CTkEntry(self.frame_derecho, placeholder_text="Cadena a Buscar", width=400)
        self.punto_buscarpropuesto.place(x=480, y=640)
        self.punto_buscarpropuesto.bind("<KeyRelease>", self.filtrar_tabla)
        
        # Configurar encabezados y columnas
        for col in columnas:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150, anchor="center")  # puedes ajustar ancho
        self.tree.heading("id", text="ID")
        self.tree.heading("nombre_estacion", text="Nombre de la Estación")
        self.tree.heading("latitud", text="Latitud")
        self.tree.heading("longitud", text="Longitud")
        self.tree.heading("altura_torre", text="Torre")
        self.tree.heading("estado", text="Estado")

        self.tree.column("id", width=30, anchor="center")
        self.tree.column("nombre_estacion", width=250, anchor="w")
        self.tree.column("latitud", width=110, anchor="center")
        self.tree.column("longitud", width=110, anchor="center")
        self.tree.column("altura_torre", width=40, anchor="center")
        self.tree.column("estado", width=110, anchor="center")
        self.tree.place(x=400, y=35, width=680, height=600)
        self.tree.bind("<<TreeviewSelect>>", self.cargar_seleccion)

        # CREA TABLA SI NO EXISTE
        crear_tabla()
        self.cargar_datos()

        # AMPLIACION DE VENTANA 
        self.after(0, lambda: self.state('zoomed'))
        self.overrideredirect(True)

    # FUNCIONES DEL MODULO

    def mostrar_messagebox_oscuro(self, titulo, mensaje):
        ventana = customtkinter.CTkToplevel()
        ventana.geometry("300x150")
        ventana.title(titulo)
        ventana.configure(fg_color="#434564")  # fondo oscuro
        etiqueta = customtkinter.CTkLabel(ventana, text=mensaje, text_color="white")
        etiqueta.pack(pady=20)
        boton = customtkinter.CTkButton(ventana, text="OK", command=ventana.destroy, fg_color="#4c607c")
        boton.pack(pady=10)
        ventana.grab_set()  # hace modal la ventana
        ventana.overrideredirect(True)

    def mostrar_messagebox_oscuro_largo(self, titulo, mensaje):
        ventana = customtkinter.CTkToplevel()
        ventana.geometry("500x250")
        ventana.title(titulo)
        ventana.configure(fg_color="#434564")  # fondo oscuro
        etiqueta = customtkinter.CTkLabel(ventana, text=mensaje, text_color="white")
        etiqueta.pack(pady=20)
        boton = customtkinter.CTkButton(ventana, text="OK", command=ventana.destroy, fg_color="#4c607c")
        boton.pack(pady=10)
        ventana.grab_set()  # hace modal la ventana
        ventana.wait_window()

    def limpiar_campos(self):
        for entry in self.campos.values():
            entry.delete(0, "end")
        self.punto_buscarpropuesto.delete(0,"end")
        self.filtrar_tabla()

    def cargar_datos(self):
        for fila in self.tree.get_children():
            self.tree.delete(fila)
        datos = obtener_todas()
        for fila in datos:
            self.tree.insert("", "end", values=fila)

    def cargar_seleccion(self, event):
        seleccionado = self.tree.selection()
        if seleccionado:
            item = self.tree.item(seleccionado[0])
            valores = item["values"]
            keys = list(self.campos.keys())
            for i, key in enumerate(keys):
                self.campos[key].delete(0, "end")
                self.campos[key].insert(0, valores[i+1])  # +1 porque el id está en la posición 0

    def agregar(self):
        try:
            datos = (
                self.campos["Nombre estación"].get().upper(),
                float(self.campos["Latitud"].get()),
                float(self.campos["Longitud"].get()),
                float(self.campos["Altura torre"].get()),
                self.campos["Estado"].get().upper(),
                self.campos["Municipio"].get().upper(),
                self.campos["Parroquia"].get().upper(),
                self.campos["Tipo estación"].get().upper(),
                self.campos["Observaciones"].get().upper()
            )
        except ValueError:
            self.mostrar_messagebox_oscuro("ERROR", "Latitud, Longitud y Altura Torre deben ser números.")
            return

        if not datos[0]:
            self.mostrar_messagebox_oscuro("ERROR", "El nombre de la estación es obligatorio.")
            return

        resultado = buscar_estacion(datos)
        
        if (resultado):
            self.mostrar_messagebox_oscuro("ERROR", "Estación ya existe en la BD")
        else: 
            insertar_estacion(datos)
            self.mostrar_messagebox_oscuro("EXITO", "Registro cargado correctamente")
            self.cargar_datos()
        self.limpiar_campos()

    def actualizar(self):
        seleccionado = self.tree.selection()
        if not seleccionado:
            self.mostrar_messagebox_oscuro("ERROR", "Selecciona una fila para actualizar.")
            return
        id_ = self.tree.item(seleccionado[0])["values"][0]
        try:
            datos = (
                self.campos["Nombre estación"].get().upper(),
                float(self.campos["Latitud"].get()),
                float(self.campos["Longitud"].get()),
                float(self.campos["Altura torre"].get()),
                self.campos["Estado"].get().upper(),
                self.campos["Municipio"].get().upper(),
                self.campos["Parroquia"].get().upper(),
                self.campos["Tipo estación"].get().upper(),
                self.campos["Observaciones"].get().upper()
            )
        except ValueError:
            self.mostrar_messagebox_oscuro("ERROR", "Latitud, Longitud y Altura Torre deben ser números.")
            return

        if not datos[0]:
            self.mostrar_messagebox_oscuro("ERROR", "El nombre de la estación es obligatorio.")
            return

        actualizar_estacion(id_, datos)
        self.mostrar_messagebox_oscuro("EXITO", "Registro actualizado correctamente")
        self.cargar_datos()
        self.limpiar_campos()

    def eliminar(self):
        seleccionado = self.tree.selection()
        if not seleccionado:
            self.mostrar_messagebox_oscuro("ERROR", "Selecciona una fila para eliminar.")
            return
        id_ = self.tree.item(seleccionado[0])["values"][0]
        # if messagebox.askyesno("Confirmar", "¿Seguro que quieres eliminar esta estación?"):
        #     eliminar_estacion(id_)
        #     self.cargar_datos()
        #     self.limpiar_campos()

        dialogo = ConfirmDialog(self, "Confirmar", "¿Seguro que quieres eliminar esta estación?")
        if dialogo.resultado:
            eliminar_estacion(id_)
            self.cargar_datos()
            self.limpiar_campos()

    def filtrar_tabla(self, event=None):
        filtro = self.punto_buscarpropuesto.get().lower()
        # Limpia la tabla
        for fila in self.tree.get_children():
            self.tree.delete(fila)
        # Obtén todos los datos de la BD
        datos = obtener_todas()
        for fila in datos:
            # Puedes ajustar aquí los campos a filtrar, por ejemplo en nombre_estacion y estado
            # fila: (id, nombre_estacion, latitud, longitud, altura_torre, estado, ...)
            if (filtro in str(fila[1]).lower() or
                filtro in str(fila[2]).lower() or
                filtro in str(fila[3]).lower() or
                filtro in str(fila[4]).lower() or
                filtro in str(fila[5]).lower()):
                self.tree.insert("", "end", values=fila)

    def exportar_a_excel(self):
        # Pedir ruta y nombre de archivo para guardar
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")],
            title="Guardar archivo Excel",
            parent=self
        )
        if not archivo:
            return  # si el usuario cancela

        # Crear libro y hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estaciones"

        # Encabezados personalizados (puedes cambiar estos textos)
        encabezados = [
            "Nombre de la Estacion",
            "Latitud",
            "Longitud",
            "Altura de la Torre",
            "Estado",
            "Municipio",
            "Parroquia",
            "Tipo de estacion",
            "Observaciones"
        ]

        # Escribir encabezados con formato
        for col_num, encabezado in enumerate(encabezados, 1):
            celda = ws.cell(row=1, column=col_num, value=encabezado)
            celda.font = Font(bold=True, color="FFFFFF")
            celda.alignment = Alignment(horizontal="center")
            celda.fill = openpyxl.styles.PatternFill("solid", fgColor="4F81BD")  # azul

        # Escribir datos de la tabla (Omitiendo el ID)
        for fila_num, item in enumerate(self.tree.get_children(), 2):
            valores = self.tree.item(item)["values"]
            # omitir el primer valor usando slicing valores [1:]
            for col_num, valor in enumerate(valores[1:], 1):
                ws.cell(row=fila_num, column=col_num, value=valor)

        # Ajustar ancho columnas (opcional)
        for col in ws.columns:
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        try:
            wb.save(archivo)
            self.mostrar_messagebox_oscuro("Éxito", f"Datos exportados correctamente a:\n{archivo}")
        except Exception as e:
            self.mostrar_messagebox_oscuro("Error", f"No se pudo guardar el archivo:\n{e}")

    def importar_desde_excel(self):
        self.mostrar_messagebox_oscuro_largo(
            "INFO",
            "Se requiere que el archivo a importar, tenga las columnas:\n\n"
            "Nombre de la Estacion (requerida), Latitud (requerida),\n"
            "Longitud (requerida), Altura de la Torre (opcional),\n"
            "Estado (requerido), Municipio (opcional), Parroquia (opcional),\n"
            "Tipo de estacion (opcional) y Observaciones (opcional)\n\n"
            "Nota: Las columnas indicadas como requeridas\n"
            "no pueden estar vacías, mientras que las\n"
            "marcadas como opcional sí pueden estar vacías"
        )

        self.mostrar_loader()        

        archivo = filedialog.askopenfilename(
            filetypes=[("Archivos Excel", "*.xlsx *.xls")],
            title="Seleccionar archivo Excel para importar",
            parent=self
        )
        if not archivo:
            self.ocultar_loader()
            return  # Canceló

        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
        except Exception as e:
            self.mostrar_messagebox_oscuro("ERROR", f"No se pudo abrir el archivo:\n{e}")
            self.ocultar_loader()
            return

        encabezados_esperados = [
            "Nombre de la Estacion",
            "Latitud",
            "Longitud",
            "Altura de la Torre",
            "Estado",
            "Municipio",
            "Parroquia",
            "Tipo de estacion",
            "Observaciones"
        ]

        encabezados_archivo = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        if encabezados_archivo[:len(encabezados_esperados)] != encabezados_esperados:
            self.mostrar_messagebox_oscuro(
                "ERROR",
                "El archivo no tiene la estructura esperada\nen los encabezados. Por favor verifique"
            )
            self.ocultar_loader()
            return

        filas_cargadas = 0
        filas_omitidas = 0
        errores = []


        # Recorremos filas desde la 2 (después de encabezado)
        for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
            valores = [cell.value if cell.value is not None else "" for cell in fila]
            nombre = str(valores[0]).strip() if valores[0] else ""

            # Si la primera columna está vacía, terminar la lectura (fin de datos)
            if nombre == "":
                break

            latitud = valores[1]
            longitud = valores[2]
            altura = valores[3]
            estado = str(valores[4]).strip() if valores[4] else ""
            municipio = str(valores[5]).strip() if valores[5] else ""
            parroquia = str(valores[6]).strip() if valores[6] else ""
            tipo_estacion = str(valores[7]).strip() if valores[7] else ""
            observaciones = str(valores[8]).strip() if valores[8] else ""

            # Validar campos obligatorios
            if estado == "" or latitud == "" or longitud == "":
                filas_omitidas += 1
                errores.append(f"Fila {fila[0].row}: faltan campos obligatorios.")
                continue

            # Validar que latitud, longitud y altura sean números
            try:
                latitud = float(latitud)
                longitud = float(longitud)
                altura = float(altura) if altura != "" else 0.0
            except ValueError:
                filas_omitidas += 1
                errores.append(f"Fila {fila[0].row}: latitud, longitud o altura inválidos.")
                continue
            

            datos = (
                nombre.upper(),
                latitud,
                longitud,
                altura,
                estado.upper(),
                municipio.upper(),
                parroquia.upper(),
                tipo_estacion.upper(),
                observaciones.upper()
            )
            try:                
                insertar_estacion(datos)
                filas_cargadas += 1
            except Exception as e:
                filas_omitidas += 1
                errores.append(f"Fila {fila[0].row}: error al insertar en base de datos: {e}")

        mensaje = f"Importación finalizada.\nFilas cargadas: {filas_cargadas}\nFilas omitidas: {filas_omitidas}"
        if errores:
            mensaje += "\n\nDetalles de errores:\n" + "\n".join(errores[:10])
            if len(errores) > 10:
                mensaje += f"\n... y {len(errores)-10} errores más."
        
        self.ocultar_loader()
        self.mostrar_messagebox_oscuro_largo("Resultado Importación", mensaje)
        self.limpiar_campos()


        self.cargar_datos()

    def vaciar_base_datos(self):
        dialogo = ConfirmDialog(self, "CONFIRMAR", "¿Seguro que quieres vaciar\ntoda la base de datos?")
        if dialogo.resultado:
            try:
                from util.db import vaciar_tabla_estaciones  # importa aquí para evitar problemas circulares
                vaciar_tabla_estaciones()
                self.mostrar_messagebox_oscuro("Éxito", "Base de datos se vació correctamente.")
                self.limpiar_campos()
                self.cargar_datos()  # recarga tabla para reflejar cambios
            except Exception as e:
                self.mostrar_messagebox_oscuro("Error", f"No se pudo vaciar la base de datos:\n{e}")
        else:
            pass

    def mostrar_loader(self, mensaje="Importando datos, por favor espere..."):
        self.label_loader = customtkinter.CTkLabel(self, text=mensaje, text_color="#DA114D", font=("Arial", 18))
        self.label_loader.place(x=300, y=500 )

    def ocultar_loader(self):
        self.label_loader.destroy()  # elimina el widget de la interfaz
        del self.label_loader  # opcional: elimina el atributo para evitar referencias

    def cerrar_modulo(self):
        self.grab_release()   # Liberar el grab (importante)
        self.destroy()        # Cerrar ventana secundaria
        if self.master is not None:
            self.master.deiconify()  # Mostrar ventana principal si estaba oculta
            self.master.focus_force()  # Dar foco a ventana principal

class ConfirmDialog(customtkinter.CTkToplevel):
    def __init__(self, master, titulo, mensaje):
        super().__init__(master)
        self.resultado = None
        self.title(titulo)
        self.geometry("350x150")
        self.configure(fg_color="#434564")  # fondo oscuro
        self.resizable(False, False)
        self.grab_set()  # ventana modal

        etiqueta = customtkinter.CTkLabel(self, text=mensaje, text_color="white", font=("Arial", 14))
        etiqueta.pack(pady=20, padx=20)

        frame_botones = customtkinter.CTkFrame(self, fg_color="transparent")
        frame_botones.pack(pady=10)

        btn_si = customtkinter.CTkButton(frame_botones, text="Sí", width=80, command=self._si, fg_color="#4c607c")
        btn_si.pack(side="left", padx=10)

        btn_no = customtkinter.CTkButton(frame_botones, text="No", width=80, command=self._no, fg_color="#4c607c")
        btn_no.pack(side="left", padx=10)

        # Centrar ventana sobre la ventana padre
        self.update_idletasks()
        x = master.winfo_rootx() + (master.winfo_width() // 2) - (self.winfo_width() // 2)
        y = master.winfo_rooty() + (master.winfo_height() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

        self.protocol("WM_DELETE_WINDOW", self._no)  # si cierra ventana, equivale a "No"

        self.wait_window()  # espera hasta que se cierre la ventana

    def _si(self):
        self.resultado = True
        self.destroy()

    def _no(self):
        self.resultado = False
        self.destroy()



